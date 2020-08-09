import ipaddress


from src.models.file_export.excel.excel import Excel
from napalm import get_network_driver
from openpyxl import load_workbook


def get_serial_and_os(device):
    facts = device.get_facts()
    serial = facts.get('serial_number')
    model = facts.get('model')
    hostname = facts.get('hostname')
    os_version = facts.get('os_version')
    return dict(serial=serial, model=model, hostname=hostname, os_version=os_version)



def get_int_status_summary(device):
    interfaces = device.get_interfaces()
    active = []
    inactive = []
    unused = []
    for interface, details in interfaces.items():
        admin = details.get("is_enabled")
        status = details.get("is_up")
        if admin and status:
            active.append(interface)
        if not admin and not status:
            unused.append(interface)
        if admin and not status:
            inactive.append(interface)
    return active, inactive, unused

def get_ip_address_usage(device):
    ips = device.get_interfaces_ip()
    ip_list = []
    for interface, details in ips.items():
        ip_dict = details.get('ipv4')
        ip_prefix = f'{list(ip_dict.keys())[0]}/{list(ip_dict.values())[0].get("prefix_length")}'
        ip_list.append(f'{interface: <25}{ip_prefix}')
    return ip_list

def get_device_details(device):
    active, inactive, unused = get_int_status_summary(device)
    ip_string = get_ip_address_usage(device)
    serial_os = get_serial_and_os(device)
    serial_os['ip_usage'] = ip_string
    serial_os['active_interface'] = active
    serial_os['active_interface_count'] = len(active)
    serial_os['inactive_interface'] = inactive 
    serial_os['inactive_interface_count'] = len(inactive)
    serial_os['unused_interface'] = unused
    serial_os['unused_interface_count'] = len(unused)
    serial_os['access'] = 'OK'
    return serial_os

def access_get_dev_details(ip, username, password, terminal='ssh'):
    try:
        driver = get_network_driver('ios')
        device = driver(ip, username, password, optional_args={'transport': terminal})
        device.open()
        if device:
            details = get_device_details(device)
            details['ip'] = ip
            device.close()
            return details
    except:
        pass
    print(f'Error connecting to {ip}, please check the credentials or terminal settings')
    return dict(ip=ip, access='FAIL', ip_usage='FAIL', active_interface='FAIL', inactive_interface='FAIL', unused_interface='FAIL', active_interface_count='FAIL', inactive_interface_count='FAIL', unused_interface_count='FAIL', serial='FAIL', model='FAIL', hostname='FAIL', os_version='FAIL')

def get_details_for_device_list(dev_list):
    dev_details_list = []
    for dev in dev_list:
        ip, username, password, terminal = dev.get('ip'), dev.get('username') , dev.get('password'), dev.get('terminal')
        details = access_get_dev_details(ip, username, password, terminal)
        dev_details_list.append(details)
    return dev_details_list

def get_ip_from_input_excel():
    filename = 'input_excel.xlsx'
    wb = load_workbook(filename = filename)
    sheet = wb[wb.sheetnames[0]]
    ip_list = []
    for row in sheet.rows:
        ip = row[0].value
        username = row[1].value
        password = row[2].value
        terminal = row[3].value
        if not ip == 'IP Address':
            ip_list.append(dict(ip=ip, username=username, password=password, terminal=terminal))
    return ip_list


def export_to_excel(details):
    updated_list = []
    for item in details:
        item_dict = {}
        for key, value in item.items():
            if isinstance(value, list):
                print(f' found list in {value}')
                value_str = ''
                for val in value:
                    value_str += f'{val},'
                value = value_str
            item_dict[key] = value
        updated_list.append(item_dict)
    excel_out = Excel.dynamic_excel(updated_list, 'Report')
    excel_out.save(filename='excel_output.xlsx')


dev_list = get_ip_from_input_excel()

details = get_details_for_device_list(dev_list)

export_to_excel(details)




