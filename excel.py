from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import csv


class Excel(object):
    
    @staticmethod
    def dynamic_excel(data, title):
        wb = Workbook()
        ws = wb.active
        ws.title = title
        row = 1
        # data_list = key_list
        key_list = data[0].keys()
        Excel.fill_row(ws, row, key_list)
        row += 1
        for column in data:
            column_list = column.values()
            Excel.fill_row(ws, row, column_list)
            row += 1
        return wb

    @staticmethod
    def dynamic_recursive_excel(data, title):
        pass

    @staticmethod
    def fill_row(work_sheet, row, data_list):
        for enum, data in enumerate(data_list, 1):
            work_sheet.cell(column=enum, row=row, value=data)

